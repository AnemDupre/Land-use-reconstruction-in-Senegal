# -*- coding: utf-8 -*-
"""
Created on Tue Jun 18 16:07:36 2024

@author: anem
This file contains functions to save and load 
the data obtained from sampling not to have to
do calculations multiple times.
"""

#imports
from datetime import date
import pandas as pd
import openpyxl


#functions
def save_sampling(params_list, path_results,
                  seed, modification_marker=None,
                  msd_list=None):
    """
    Creates an xlsx file containing the parameter values
    and associated MSDs.

    Parameters
    ----------
    msd_list : list of MSD values
            obtained by comparing the output 
            of the model for parameter values 
            at the same index in params_list
            and FAO land use data.
    params_list : list containg dictionaries 
            of parameter values sampled for
            parameter optimization.
    path_results : str path where we want to
            save msd and parameter values.

    Returns
    -------
    path_2results : STR PATH OF FILE IN WHICH
            MSD_LIST AND PARAMS_LIST WILL BE
            SAVED.

    """
    cur_date = str(date.today()) #fetch current date
    numb_samples = len(params_list)
    
    
    #determining name
    if modification_marker!=None:
        path_2results = path_results + f'{cur_date}_{numb_samples}samples_seed_{seed}_{modification_marker}_params_msds.xlsx'
    else:
        path_2results = path_results + f'{cur_date}_{numb_samples}samples_seed_{seed}_params_msds.xlsx'
    
    #saving results
    df = pd.DataFrame.from_dict(params_list)
    df.to_excel(path_2results)
    
    if msd_list!=None:
        #adding msd values
        column = len(params_list[0].keys())+2 #at which column to put msd values
        workbook = openpyxl.load_workbook(path_2results)
        worksheet = workbook.worksheets[0]
        cell_title = worksheet.cell(row=1, column=column)
        cell_title.value = 'MSD'
        y=2
        for x in range(len(msd_list)):
            cell_to_write = worksheet.cell(row=y, column=column)
            cell_to_write.value = msd_list[x]
            y += 1
    workbook.save(path_2results)
    
    return path_2results



def save_outputs(path_results, seed, crop_df, past_df, crop_subs_df, crop_mark_df, fal_df, un_df, veg_df, intensification_df,
                 modification_marker=None):
    cur_date = str(date.today()) #fetch current date
    numb_samples = len(crop_df.columns) -1 #one of the columns coreesponds to the years and not to a simulation
    
    categories = ["crop", "past", "crop_subs", "crop_mark", "fal", "un", "veg", "intensification"]
    list_outputs = [crop_df, past_df, crop_subs_df, crop_mark_df, fal_df, un_df, veg_df, intensification_df]
    
    for category, output in zip(categories, list_outputs):
        #path name
        if modification_marker!=None:
            path_2results = path_results + f'{cur_date}_{numb_samples}samples_seed_{seed}_{modification_marker}_{category}.pkl'
        else:
            path_2results = path_results + f'{cur_date}_{numb_samples}samples_seed_{seed}_{category}.pkl'
        #saving output
        output.to_pickle(path_2results)



def load_saved_params_msds(path_results, header, modification_marker=None):
    """
    Fetches the data obtained from sampling and 
    saved in path with the function save_sampling().

    Parameters
    ----------
    path : STR PATH OF FILE IN WHICH
        MSD_LIST AND PARAMS_LIST WILL BE
        SAVED.

    Returns
    -------
    msd_list : LIST OF MSD VALUES OBTAINED
        BY COMPARING THE OUTPUT OF THE MODEL
        FOR PARAMETER VALUES AT THE SAME
        INDEX IN PARAMS_LIST AND FAO LAND 
        USE DATA.
    params_list : LIST CONTAINING DICTIONARIES 
        OF PARAMETER VALUES SAMPLED FOR 
        PARAMETER OPTIMIZATION.

    """
    if modification_marker!=None:
        path = path_results + header + f"_{modification_marker}_params_msds.xlsx"
    else:
        path = path_results + header + "_params_msds.xlsx"
    
    # Load the Excel file
    file = pd.read_excel(path)
    
    # Retrieve params_list from the DataFrame
    params_list = file.to_dict(orient='records')
    # Iterate over each dictionary
    for d in params_list:
        # Check if the "msd" key exists in the dictionary
        if "MSD" in d:
            # Remove the "msd" key and its associated value
            d.pop("MSD")
        if 'Unnamed: 0' in d:
            d.pop("Unnamed: 0")
    if "MSD" in file.columns:
        # Retrieve msd_list from the 'MSD' column of the DataFrame
        msd_list = file['MSD'].tolist()
    
    if "msd_list" in locals():
        return msd_list, params_list
    else :
        return params_list



def load_saved_outputs(path_results, file_header, modification_marker=None):
    categories = ["crop", "past", "crop_subs", "crop_mark", "fal", "un", "veg", "intensification"]
    list_outputs = []
    
    for category in categories:
        if modification_marker!=None:
            path = path_results + file_header + f"_{modification_marker}_{category}.pkl"
        else:
            path = path_results + file_header + f"_{category}.pkl"
        
        df = pd.read_pickle(path)
        list_outputs.append(df)
        
    [crop_df, past_df, crop_subs_df, crop_mark_df, fal_df, un_df, veg_df, intensification_df] = list_outputs
    
    return crop_df, past_df, crop_subs_df, crop_mark_df, fal_df, un_df, veg_df, intensification_df